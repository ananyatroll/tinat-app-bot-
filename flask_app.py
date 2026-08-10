"""
Tinat Telegram bot - Python/Flask migration for PythonAnywhere Free Tier.

The bot runs entirely inside the request/response cycle of a Flask web worker.
Telegram is configured in webhook mode and POSTs every update to
https://<username>.pythonanywhere.com/telegram-webhook

Feature parity with the current Node.js bot:
  - /start package chooser (EUEE Preo, Freshman, UAT, University Department, Exit Exam)
  - state machine: package -> phone (verified via Telegram contact share) ->
    payment method (CBE/Telebirr) -> name -> transaction link -> transaction ID
  - admin approval with Approve / Reject inline buttons
  - Approve immediately reserves one unused voucher from the correct package
    pool and sends the phrase to the buyer; Reject marks the payment rejected
  - support role (/pending or the Assign buttons) for assigning vouchers to
    already-approved purchases that are missing one
  - voucher assignment tracked in data/vouchers.json with owner / phone /
    transaction / assigned-by audit fields
  - Excel export of approved requests (one sheet per package) sent to the admin
  - redemption API that verifies Telegram Mini App initData, voucher ownership,
    payment approval and voucher state, then atomically marks the voucher
    REDEEMED and writes the entitlement to data/entitlements.json
"""

import base64
import copy
import hashlib
import hmac
import json
import logging
import os
import re
import secrets
import time
from contextlib import contextmanager
from datetime import datetime, timezone
from urllib.parse import unquote

import requests
from flask import Flask, jsonify, request
from openpyxl import Workbook
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Configuration / environment
# ---------------------------------------------------------------------------


def _load_dotenv(path=None):
    """Tiny .env loader (no extra dependency) so local testing works with
    the existing .env file. Real env vars always win."""
    path = path or os.path.join(BASE_DIR, '.env')
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue
                key, _, value = line.partition('=')
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except OSError:
        pass


def _abs_path(value, default):
    """Resolve a possibly-relative configured path against BASE_DIR so the
    app works regardless of the process working directory (important on
    PythonAnywhere, where the WSGI working directory is configurable)."""
    value = (value or '').strip() or default
    if not os.path.isabs(value):
        value = os.path.join(BASE_DIR, value)
    return os.path.normpath(value)


_load_dotenv()

TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN') or os.environ.get('BOT_TOKEN', '')
ADMIN_CHAT_ID = os.environ.get('ADMIN_CHAT_ID', '')
SUPPORT_CHAT_ID = os.environ.get('SUPPORT_CHAT_ID') or ADMIN_CHAT_ID or ''
DATA_DIR = _abs_path(os.environ.get('DATA_DIR'), os.path.join(BASE_DIR, 'data'))
PHRASES_DIR = _abs_path(os.environ.get('PHRASES_DIR'), os.path.join(BASE_DIR, 'phrases'))
EXPORTS_DIR = _abs_path(os.environ.get('EXPORTS_DIR'), os.path.join(BASE_DIR, 'exports'))
USERS_FILE = _abs_path(os.environ.get('USERS_FILE'), os.path.join(DATA_DIR, 'users.json'))
VOUCHERS_FILE = _abs_path(os.environ.get('VOUCHERS_FILE'), os.path.join(DATA_DIR, 'vouchers.json'))
ENTITLEMENTS_FILE = _abs_path(os.environ.get('ENTITLEMENTS_FILE'), os.path.join(DATA_DIR, 'entitlements.json'))
PROCESSED_UPDATES_FILE = _abs_path(os.environ.get('PROCESSED_UPDATES_FILE'), os.path.join(DATA_DIR, 'processed_updates.json'))
MAX_RECENT_UPDATES = 500
UPDATE_DEDUPE_TTL_SECONDS = 60 * 60 * 24

REDEEM_RATE_LIMIT_MAX = int(os.environ.get('REDEEM_RATE_LIMIT_MAX', '5'))
REDEEM_RATE_LIMIT_WINDOW_MS = int(os.environ.get('REDEEM_RATE_LIMIT_WINDOW_MS', str(10 * 60 * 1000)))
INIT_DATA_MAX_AGE_SECONDS = int(os.environ.get('INIT_DATA_MAX_AGE_SECONDS', str(24 * 60 * 60)))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s: %(message)s',
)
logger = logging.getLogger('tinat-bot')

if not TELEGRAM_BOT_TOKEN:
    logger.warning('TELEGRAM_BOT_TOKEN is not set. The bot will not process messages.')
if not ADMIN_CHAT_ID:
    logger.warning('ADMIN_CHAT_ID is not set. The approval flow will not work.')


# ---------------------------------------------------------------------------
# Packages
# ---------------------------------------------------------------------------

DEFAULT_PACKAGES = [
    {'key': 'euee-preo', 'label': 'EUEE Preo', 'priceCents': 25000, 'currency': 'ETB', 'phrasePool': 'euee-preo'},
    {'key': 'freshman', 'label': 'Freshman', 'priceCents': 25000, 'currency': 'ETB', 'phrasePool': 'freshman'},
    {'key': 'uat', 'label': 'UAT', 'priceCents': 25000, 'currency': 'ETB', 'phrasePool': 'uat'},
    {'key': 'university-department', 'label': 'University Department', 'priceCents': 25000, 'currency': 'ETB', 'phrasePool': 'university-department'},
    {'key': 'exit-exam', 'label': 'Exit Exam', 'priceCents': 25000, 'currency': 'ETB', 'phrasePool': 'exit-exam'},
]


def load_packages():
    raw = os.environ.get('PACKAGES_JSON')
    if not raw:
        return copy.deepcopy(DEFAULT_PACKAGES)

    try:
        parsed = json.loads(raw)
    except ValueError:
        raise ValueError('PACKAGES_JSON must be valid JSON')

    if not isinstance(parsed, list) or not parsed:
        raise ValueError('PACKAGES_JSON must contain at least one package')

    packages = []
    for index, item in enumerate(parsed):
        key = str(item.get('key') or item.get('id') or 'package_%d' % (index + 1)).strip()
        label = str(item.get('label') or item.get('name') or 'Package %d' % (index + 1)).strip()
        try:
            price = int(float(item.get('priceCents', item.get('price_cents', 0))))
        except (TypeError, ValueError):
            price = 0
        currency = str(item.get('currency') or os.environ.get('CURRENCY', 'USD')).upper()
        pool = str(item.get('phrasePool') or item.get('pool') or item.get('key') or item.get('id') or key).strip()
        if key and label and price > 0:
            packages.append({
                'key': key,
                'label': label,
                'priceCents': price,
                'currency': currency,
                'phrasePool': pool,
            })
    return packages


PACKAGES = load_packages()


def get_package_by_key(package_key):
    for pkg in PACKAGES:
        if pkg['key'] == package_key:
            return pkg
    return None


def format_money(cents, currency='USD'):
    return '%.2f %s' % (cents / 100, currency)


def build_package_keyboard():
    rows = [[
        {'text': '%s - %s' % (pkg['label'], format_money(pkg['priceCents'], pkg['currency'])),
         'callback_data': 'package:%s' % pkg['key']}
    ] for pkg in PACKAGES]
    return {'inline_keyboard': rows}


def get_start_message():
    lines = [
        'Welcome to Tinat.',
        'Pick a package below, then I will ask you to share your phone number, payment method, name, transaction link, and transaction ID.',
        '',
    ]
    lines += ['%s - %s' % (pkg['label'], format_money(pkg['priceCents'], pkg['currency'])) for pkg in PACKAGES]
    lines += ['', 'After approval, our support team sends you one secret phrase for your selected package.']
    return '\n'.join(lines)


def get_package_selection_message():
    lines = [
        'Choose your package first:',
    ]
    lines += ['%s - %s' % (pkg['label'], format_money(pkg['priceCents'], pkg['currency'])) for pkg in PACKAGES]
    lines += ['', 'I will then ask you to share your phone number, payment method, name, transaction link, and transaction ID.']
    return '\n'.join(lines)


# ---------------------------------------------------------------------------
# Payment methods
# ---------------------------------------------------------------------------

PAYMENT_METHODS = [
    {'key': 'cbe', 'label': 'Commercial Bank of Ethiopia (CBE)'},
    {'key': 'telebirr', 'label': 'Telebirr'},
]

PAYMENT_LINK_PATTERNS = {
    'cbe': re.compile(r'^https://mbreciept\.cbe\.com\.et/\S+', re.IGNORECASE),
    'telebirr': re.compile(r'^https://transactioninfo\.ethiotelecom\.et/receipt/\S+', re.IGNORECASE),
}

PAYMENT_ID_PATTERNS = {
    'cbe': re.compile(r'^FT\d{4,}[A-Z0-9]{3,}$', re.IGNORECASE),
    'telebirr': re.compile(r'^[A-Z]{3}\d{1,}[A-Z0-9]{2,}$', re.IGNORECASE),
}


def is_valid_transaction_link(method_key, link):
    pattern = PAYMENT_LINK_PATTERNS.get(method_key)
    return bool(pattern and pattern.match(str(link or '').strip()))


def is_valid_transaction_id(method_key, value):
    pattern = PAYMENT_ID_PATTERNS.get(method_key)
    return bool(pattern and pattern.match(str(value or '').strip()))


def link_format_hint(method_key):
    if method_key == 'cbe':
        return 'https://mbreciept.cbe.com.et/<your-receipt-code>'
    if method_key == 'telebirr':
        return 'https://transactioninfo.ethiotelecom.et/receipt/<your-transaction-id>'
    return 'the full receipt link'


def id_format_hint(method_key):
    if method_key == 'cbe':
        return 'starts with FT, e.g. FT26222QKMBG'
    if method_key == 'telebirr':
        return 'letters and digits, e.g. DGJ22CMPJM'
    return 'your transaction ID'


def get_payment_method(method_key):
    for method in PAYMENT_METHODS:
        if method['key'] == method_key:
            return method
    return None


def build_payment_method_keyboard():
    rows = [[
        {'text': method['label'], 'callback_data': 'method:%s' % method['key']}
    ] for method in PAYMENT_METHODS]
    return {'inline_keyboard': rows}


def build_phone_share_keyboard():
    return {
        'keyboard': [[{'text': 'Share Phone Number', 'request_contact': True}]],
        'resize_keyboard': True,
        'one_time_keyboard': True,
    }


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------


def utcnow():
    return datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')


def random_id(length):
    raw = base64.urlsafe_b64encode(secrets.token_bytes(length)).decode('ascii').rstrip('=')
    return raw[:length]


def is_plausible_phone(value):
    digits = re.sub(r'\D', '', str(value or ''))
    return 9 <= len(digits) <= 15


def _acquire_lock(fh):
    if os.name == 'nt':
        try:
            import msvcrt
            fh.seek(0)
            if not fh.read(1):
                fh.write(b'\0')
                fh.flush()
            fh.seek(0)
            msvcrt.locking(fh.fileno(), msvcrt.LK_LOCK, 1)
        except ImportError:
            pass
    else:
        try:
            import fcntl
            fcntl.flock(fh.fileno(), fcntl.LOCK_EX)
        except ImportError:
            pass


def _release_lock(fh):
    if os.name == 'nt':
        try:
            import msvcrt
            fh.seek(0)
            msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, 1)
        except Exception:
            pass
    else:
        try:
            import fcntl
            fcntl.flock(fh.fileno(), fcntl.LOCK_UN)
        except Exception:
            pass


@contextmanager
def file_locked(path):
    """Cross-platform advisory file lock (fcntl on Linux/PythonAnywhere,
    msvcrt fallback on Windows for local testing)."""
    lock_path = path + '.lock'
    os.makedirs(os.path.dirname(lock_path) or '.', exist_ok=True)
    fh = open(lock_path, 'a+b')
    try:
        _acquire_lock(fh)
        yield
    finally:
        _release_lock(fh)
        fh.close()


def default_users():
    return {'users': {}, 'requests': {}, 'drafts': {}}


def default_vouchers():
    return {'packages': {}, 'issued': {}}


def default_entitlements():
    return {'entitlements': {}}


def read_json(path, default):
    with file_locked(path):
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                return json.load(fh)
        except (OSError, ValueError):
            return copy.deepcopy(default)


def mutate_json(path, default, fn):
    """Atomically read -> mutate -> write a JSON file under a file lock.

    fn(data) may return anything; it is returned unchanged."""
    with file_locked(path):
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
        except (OSError, ValueError):
            data = copy.deepcopy(default)
        result = fn(data)
        os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
        tmp_path = path + '.tmp'
        with open(tmp_path, 'w', encoding='utf-8') as fh:
            json.dump(data, fh, indent=2, ensure_ascii=False)
        os.replace(tmp_path, path)
        return result


def record_update(update_id):
    """Webhook delivery is at-least-once: Telegram retries a POST until it gets
    a 2xx, which would otherwise process the same update twice (advancing a
    draft two steps or double-submitting a request). Each update carries a
    monotonic update_id, so remember recent ones and skip repeats."""
    if update_id is None:
        return False

    def _record(store):
        key = str(update_id)
        if key in store:
            return True

        now = utcnow()
        store[key] = now

        cutoff = datetime.now(timezone.utc).timestamp() - UPDATE_DEDUPE_TTL_SECONDS
        for seen_key, seen_at in list(store.items()):
            try:
                ts = datetime.fromisoformat(str(seen_at).replace('Z', '+00:00')).timestamp()
            except ValueError:
                ts = 0
            if ts < cutoff:
                store.pop(seen_key, None)

        while len(store) > MAX_RECENT_UPDATES:
            store.pop(min(store, key=lambda k: store[k]), None)

        return False

    return mutate_json(PROCESSED_UPDATES_FILE, {}, _record)


# ---------------------------------------------------------------------------
# Telegram API calls
# ---------------------------------------------------------------------------


def tg_call(method, timeout=30, **kwargs):
    if not TELEGRAM_BOT_TOKEN:
        return None
    url = 'https://api.telegram.org/bot%s/%s' % (TELEGRAM_BOT_TOKEN, method)
    try:
        response = requests.post(url, timeout=timeout, **kwargs)
        payload = response.json()
        if not payload.get('ok'):
            logger.warning('Telegram %s error: %s', method, payload)
        return payload
    except requests.RequestException as exc:
        logger.warning('Telegram %s network error: %s', method, exc)
        return None


def send_message(chat_id, text, reply_markup=None):
    params = {'chat_id': chat_id, 'text': text}
    if reply_markup is not None:
        params['reply_markup'] = reply_markup
    return tg_call('sendMessage', json=params)


def answer_callback_query(callback_query_id, text=None):
    params = {'callback_query_id': callback_query_id}
    if text:
        params['text'] = text
    return tg_call('answerCallbackQuery', json=params)


def edit_message_reply_markup(chat_id, message_id):
    params = {
        'chat_id': chat_id,
        'message_id': message_id,
        'reply_markup': {'inline_keyboard': []},
    }
    return tg_call('editMessageReplyMarkup', json=params)


def send_document(chat_id, file_path, caption=None):
    data = {'chat_id': chat_id}
    if caption:
        data['caption'] = caption
    with open(file_path, 'rb') as fh:
        files = {
            'document': (os.path.basename(file_path), fh,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        return tg_call('sendDocument', data=data, files=files)


# ---------------------------------------------------------------------------
# Roles
# ---------------------------------------------------------------------------


def is_admin(user_id):
    return str(user_id) == str(ADMIN_CHAT_ID)


def is_support(user_id):
    return bool(SUPPORT_CHAT_ID) and str(user_id) == str(SUPPORT_CHAT_ID)


def can_assign_vouchers(user_id):
    return is_admin(user_id) or is_support(user_id)


# ---------------------------------------------------------------------------
# Message builders
# ---------------------------------------------------------------------------


def build_request_message(request):
    user = request.get('user') or {}
    phone = request.get('phone') or {}
    phone_text = phone.get('number') or 'n/a'
    phone_verified = ' (verified)' if phone.get('verified') else ' (unverified)'
    full_name = ('%s %s' % (user.get('firstName') or '', user.get('lastName') or '')).strip()
    return '\n'.join([
        'New access request pending review:',
        'Request ID: %s' % request.get('requestId'),
        'Package: %s (%s)' % (request.get('packageLabel'), format_money(request.get('priceCents'), request.get('currency'))),
        'Payment Method: %s' % (request.get('paymentMethodLabel') or request.get('paymentMethod') or 'N/A'),
        'User: %s' % full_name,
        'Telegram: @%s (%s)' % (user.get('username') or 'no_username', user.get('id')),
        'Phone: %s%s' % (phone_text, phone_verified),
        'Transaction ID: %s' % request.get('transactionId'),
        'Transaction Link: %s' % request.get('transactionLink'),
        '',
        'Approve this request only after verifying the payment.',
    ])


def build_pending_message(request):
    return '\n'.join([
        'Your payment details were submitted.',
        'Request ID: %s' % request.get('requestId'),
        'I sent it to the admin for verification.',
        'You will get your voucher phrase after approval.',
    ])


def build_approved_pending_voucher_message(request):
    return '\n'.join([
        'Your payment was approved.',
        'Our support team is assigning your voucher.',
        'You will receive your voucher phrase in this chat shortly.',
        'Request ID: %s' % request.get('requestId'),
    ])


def build_rejected_message(request):
    return '\n'.join([
        'Your payment was not approved.',
        'The admin could not verify the payment details.',
        'If you believe this is a mistake, submit a new request with /buy.',
        'Request ID: %s' % request.get('requestId'),
    ])


def build_approved_message(package_label, voucher_phrase):
    return '\n'.join([
        'Payment approved. Your Tinat voucher phrase is below:',
        'Package: %s' % package_label,
        'Phrase: %s' % voucher_phrase,
        '',
        'Keep this phrase safe. If you lose it, send /myaccess in this bot.',
    ])


def build_pending_assignment_message(pending):
    lines = ['Approved purchases waiting for voucher assignment:']
    if not pending:
        lines.append('(none)')
    for index, req in enumerate(pending, start=1):
        user = req.get('user') or {}
        phone = (req.get('phone') or {}).get('number') or 'n/a'
        verified = ' (verified)' if (req.get('phone') or {}).get('verified') else ' (unverified)'
        full_name = ('%s %s' % (user.get('firstName') or '', user.get('lastName') or '')).strip()
        lines.append('%d. %s | %s | %s | @%s | %s%s'
                     % (index, req.get('requestId'), req.get('packageLabel'), full_name,
                        user.get('username') or 'no_username', phone, verified))
    return '\n'.join(lines)


def notify_admin(request):
    if not ADMIN_CHAT_ID:
        raise RuntimeError('Missing ADMIN_CHAT_ID in environment')
    send_message(ADMIN_CHAT_ID, build_request_message(request), reply_markup={
        'inline_keyboard': [[
            {'text': 'Approve', 'callback_data': 'approve:%s' % request['requestId']},
            {'text': 'Reject', 'callback_data': 'reject:%s' % request['requestId']},
        ]]
    })


# ---------------------------------------------------------------------------
# Voucher pools
# ---------------------------------------------------------------------------


class PoolEmptyError(Exception):
    pass


def _load_pool_phrases(pool_key):
    """Read phrases for a pool from a text file inside PHRASES_DIR.

    File layout:
        <PHRASES_DIR>/<pool_key>.txt        -> phrase per line, ignores blanks
        <PHRASES_DIR>/<pool_key>@<suffix>.txt -> lines starting with '<phrase>:'
                                               are split into phrase / extra
    """
    directory = PHRASES_DIR
    phrase_file = os.path.join(directory, '%s.txt' % pool_key)
    if os.path.exists(phrase_file):
        with open(phrase_file, 'r', encoding='utf-8') as fh:
            lines = [line.strip() for line in fh if line.strip()]
        return [line for line in lines if not line.startswith('#')]

    entries = []
    if os.path.isdir(directory):
        prefix = '%s@' % pool_key
        for name in os.listdir(directory):
            if not name.startswith(prefix) or not name.endswith('.txt'):
                continue
            with open(os.path.join(directory, name), 'r', encoding='utf-8') as fh:
                for line in fh:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    if ':' in line:
                        phrase, _, extra = line.partition(':')
                        entries.append({'phrase': phrase.strip(), 'extra': extra.strip()})
                    else:
                        entries.append({'phrase': line, 'extra': ''})
    return entries


def _as_pool_entry(item):
    if isinstance(item, dict):
        return {'phrase': str(item.get('phrase') or item.get('code') or ''), 'extra': str(item.get('extra') or '')}
    return {'phrase': str(item), 'extra': ''}


def _normalize_pool(pool):
    """Migrate older pool layouts (available/issued, plain-string lists) into
    the current unused/assigned dict layout so reserved phrases are never
    double-issued when old data is reused."""
    if 'unused' not in pool:
        pool['unused'] = []
        old_available = pool.pop('available', None)
        if old_available:
            pool['unused'] = [_as_pool_entry(item) for item in old_available]
    if 'assigned' not in pool:
        pool['assigned'] = []
        old_issued = pool.pop('issued', None)
        if isinstance(old_issued, dict):
            pool['assigned'] = [_as_pool_entry(item) for item in old_issued.values()]
        elif isinstance(old_issued, list):
            pool['assigned'] = [_as_pool_entry(item) for item in old_issued]
    return pool


def _hydrate_pool(data, pool_key):
    """Ensure 'data' has a complete entry for pool_key. Loads every phrase that
    is not already tracked in the store so new phrase files show up in
    /assign without a restart."""
    pool = data['packages'].get(pool_key)
    if pool is None:
        pool = {'unused': [], 'assigned': []}
        data['packages'][pool_key] = pool
    pool = _normalize_pool(pool)

    unused_by_phrase = {str(item['phrase']) for item in pool['unused']}
    assigned_by_phrase = {str(item['phrase']) for item in pool['assigned']}

    loaded = _load_pool_phrases(pool_key)
    for entry in loaded:
        phrase = entry if isinstance(entry, str) else entry.get('phrase')
        extra = '' if isinstance(entry, str) else entry.get('extra', '')
        key = str(phrase)
        if key in unused_by_phrase or key in assigned_by_phrase:
            continue
        pool['unused'].append({'phrase': key, 'extra': extra})

    return pool


def reserve_voucher(pool_key):
    """Pick and remove the first unused voucher for a pool atomically.
    Returns {'phrase': ..., 'extra': ...} or raises PoolEmptyError."""
    def _reserve(data):
        pool = _hydrate_pool(data, pool_key)
        if not pool['unused']:
            raise PoolEmptyError('Pool "%s" is empty' % pool_key)
        entry = pool['unused'].pop(0)
        assigned = pool.setdefault('assigned', [])
        assigned.append(dict(entry))
        return dict(entry)

    return mutate_json(VOUCHERS_FILE, default_vouchers(), _reserve)


def record_voucher_assignment(entry):
    """Persist a fully-assigned voucher (with owner details) into the 'issued'
    map so get_voucher_for_user / find_voucher / redeem_code can find it."""
    def _record(data):
        issued = data.setdefault('issued', {})
        issued[str(entry.get('phrase'))] = entry
        return True

    return mutate_json(VOUCHERS_FILE, default_vouchers(), _record)


def get_voucher_for_user(user_id):
    """Look up the most recently assigned, not-yet-redeemed voucher of a user."""
    data = read_json(VOUCHERS_FILE, default_vouchers())
    issued = data.get('issued') or {}
    entries = []
    for entry in issued.values():
        if str(entry.get('ownerId')) != str(user_id):
            continue
        status = normalize_voucher_status(entry)
        if status in ('assigned', 'issued'):
            entries.append(entry)
    if not entries:
        return None
    entries.sort(key=lambda item: str(item.get('assignedAt') or ''), reverse=True)
    return entries[0]


def normalize_voucher_status(entry):
    if entry.get('redeemed'):
        return 'redeemed'
    if entry.get('disbursed') or entry.get('sent'):
        return 'disbursed'
    if entry.get('assigned'):
        return 'assigned'
    return 'issued'


# ---------------------------------------------------------------------------
# Redemption API helpers
# ---------------------------------------------------------------------------


def verify_init_data(init_data):
    """Validate a Telegram Mini App initData string.

    Returns the parsed user object (dict) when the data is authentic and
    within INIT_DATA_MAX_AGE_SECONDS, otherwise returns None.

    initData format (urlencoded pairs, sorted, joined with '\\n'):
        auth_date=<unix>\\nquery_id=<..>\\nuser=<urlencoded-json>...&hash=<sha256>
    The hash is the HMAC-SHA256 of that signature string using a secret key
    derived from the bot token.
    """
    if not init_data or not TELEGRAM_BOT_TOKEN:
        return None

    try:
        parsed = dict(pair.split('=', 1) for pair in str(init_data).split('&'))
    except ValueError:
        return None

    received_hash = parsed.pop('hash', None)
    if not received_hash:
        return None

    auth_date = parsed.get('auth_date')
    try:
        if abs(int(auth_date) - time.time()) > INIT_DATA_MAX_AGE_SECONDS:
            return None
    except (TypeError, ValueError):
        return None

    signature = '\n'.join('%s=%s' % (k, parsed[k]) for k in sorted(parsed))
    secret_key = hmac.new(b'WebAppData', TELEGRAM_BOT_TOKEN.encode('utf-8'), hashlib.sha256).digest()
    expected_hash = hmac.new(secret_key, signature.encode('utf-8'), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected_hash, received_hash):
        return None

    try:
        user = json.loads(unquote(parsed.get('user', '{}')))
    except ValueError:
        user = None
    return user if isinstance(user, dict) else None


RATE_LIMIT = {}


def rate_limit_redeem(user_key, max_hits=REDEEM_RATE_LIMIT_MAX, window_ms=REDEEM_RATE_LIMIT_WINDOW_MS):
    now_ms = int(time.time() * 1000)
    hits = [ts for ts in RATE_LIMIT.get(user_key, []) if now_ms - ts < window_ms]
    if len(hits) >= max_hits:
        RATE_LIMIT[user_key] = hits
        return False
    hits.append(now_ms)
    RATE_LIMIT[user_key] = hits
    return True


def find_voucher(owner_id, phrase):
    data = read_json(VOUCHERS_FILE, default_vouchers())
    issued = data.get('issued') or {}
    for entry in issued.values():
        if str(entry.get('ownerId')) != str(owner_id):
            continue
        if str(entry.get('phrase')) == str(phrase):
            return entry
    return None


def redeem_code(owner_id, phrase):
    """Atomically mark an owned voucher as redeemed and write the entitlement.
    Returns the entitlement record or a dict error."""
    def _redeem(data):
        issued = data.setdefault('issued', {})
        for key, entry in issued.items():
            if str(entry.get('ownerId')) != str(owner_id):
                continue
            if str(entry.get('phrase')) != str(phrase):
                continue
            status = normalize_voucher_status(entry)
            if status == 'redeemed':
                return {'error': 'Voucher already redeemed'}
            if status != 'assigned':
                return {'error': 'Voucher is not assigned to you yet'}

            entry['redeemed'] = True
            entry['redeemedAt'] = utcnow()
            entry['status'] = 'redeemed'

            entitlement = {
                'entitlementId': random_id(16),
                'ownerId': str(owner_id),
                'phrase': entry.get('phrase'),
                'packageKey': entry.get('packageKey'),
                'packageLabel': entry.get('packageLabel'),
                'assignedAt': entry.get('assignedAt'),
                'redeemedAt': entry['redeemedAt'],
                'phone': entry.get('phone'),
                'ownerName': entry.get('ownerName'),
            }
            return entitlement

        return {'error': 'No matching voucher found'}

    result = mutate_json(VOUCHERS_FILE, default_vouchers(), _redeem)
    if isinstance(result, dict) and 'error' in result:
        return result

    entitlement = result

    def _record(data):
        data.setdefault('entitlements', {})[entitlement['entitlementId']] = entitlement
        return True

    mutate_json(ENTITLEMENTS_FILE, default_entitlements(), _record)
    return entitlement


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


EXPORT_HEADERS = [
    'Request ID', 'Package', 'User', 'Telegram Username', 'User ID', 'Phone',
    'Payment Method', 'Transaction ID', 'Transaction Link', 'Voucher Phrase',
    'Status', 'Approved At', 'Voucher Assigned At', 'Voucher Owner ID',
    'Owner Phone', 'Owner Name',
]


def write_exports(requests_by_package):
    """Write one Excel file per package plus a totals file.

    requests_by_package: {package_key: [request, ...]}
    Voucher/assignment columns are filled from vouchers.json 'issued' entries.
    Returns a list of written file paths.
    """
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    voucher_data = read_json(VOUCHERS_FILE, default_vouchers())
    approved_map = {}
    for entry in (voucher_data.get('issued') or {}).values():
        if entry.get('requestId'):
            approved_map[entry['requestId']] = entry

    written = []
    for package_key, requests in requests_by_package.items():
        pkg = get_package_by_key(package_key)
        label = (pkg['label'] if pkg else package_key).replace(' ', '_')
        filename = '%s_%s.xlsx' % (datetime.now().strftime('%Y%m%d_%H%M%S'), label)
        file_path = os.path.join(EXPORTS_DIR, filename)
        write_excel_requests(file_path, requests, approved_map=approved_map)
        written.append(file_path)

    totals_path = os.path.join(EXPORTS_DIR,
                               '%s_all_packages.xlsx' % datetime.now().strftime('%Y%m%d_%H%M%S'))
    write_excel_totals(totals_path, requests_by_package)
    written.append(totals_path)
    return written


def write_excel_requests(file_path, requests, approved_map=None):
    """Rows: one per request. Voucher/assignment columns filled from the
    approved entries only when the request belongs to that package."""
    approved_map = approved_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Requests'
    _write_excel_headers(ws, EXPORT_HEADERS)

    for req in requests:
        voucher = approved_map.get(req.get('requestId'))
        ws.append([
            req.get('requestId'),
            req.get('packageLabel'),
            ('%s %s' % ((req.get('user') or {}).get('firstName') or '',
                        (req.get('user') or {}).get('lastName') or '')).strip(),
            (req.get('user') or {}).get('username') or '',
            (req.get('user') or {}).get('id') or '',
            (req.get('phone') or {}).get('number') or '',
            req.get('paymentMethodLabel') or '',
            req.get('transactionId') or '',
            req.get('transactionLink') or '',
            (voucher or {}).get('phrase') if voucher else '',
            normalize_voucher_status(voucher) if voucher else req.get('status') or '',
            req.get('approvedAt') or '',
            (voucher or {}).get('assignedAt') if voucher else '',
            (voucher or {}).get('ownerId') if voucher else '',
            (voucher or {}).get('phone') if voucher else '',
            (voucher or {}).get('ownerName') if voucher else '',
        ])

    _autofit_excel(ws)
    wb.save(file_path)


def write_excel_totals(file_path, requests_by_package):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    _write_excel_headers(ws, ['Package', 'Total Requests', 'Approved', 'Pending', 'Rejected'])

    for package_key, requests in requests_by_package.items():
        pkg = get_package_by_key(package_key)
        label = pkg['label'] if pkg else package_key
        approved = sum(1 for req in requests if req.get('status') == 'approved')
        pending = sum(1 for req in requests if req.get('status') in ('pending', 'processing'))
        rejected = sum(1 for req in requests if req.get('status') == 'rejected')
        ws.append([label, len(requests), approved, pending, rejected])

    _autofit_excel(ws)
    wb.save(file_path)


def _write_excel_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _autofit_excel(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value_len = len(str(cell.value))
            if value_len > widths.get(cell.column_letter, 0):
                widths[cell.column_letter] = value_len
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


# ---------------------------------------------------------------------------
# Store / user / draft helpers
# ---------------------------------------------------------------------------


def _store_user_profile(data, user):
    uid = str(user.get('id'))
    users = data.setdefault('users', {})
    users[uid] = {
        'firstName': user.get('first_name') or user.get('firstName') or '',
        'lastName': user.get('last_name') or user.get('lastName') or '',
        'username': user.get('username') or '',
        'languageCode': user.get('language_code') or '',
        'lastSeen': utcnow(),
    }


def _get_user_profile(data, user_id):
    return (data.get('users') or {}).get(str(user_id)) or {}


def _new_request(user_id, profile, package, phone):
    now = utcnow()
    return {
        'requestId': random_id(12),
        'userId': str(user_id),
        'user': {
            'id': str(user_id),
            'firstName': profile.get('firstName') or '',
            'lastName': profile.get('lastName') or '',
            'username': profile.get('username') or '',
            'languageCode': profile.get('languageCode') or '',
        },
        'packageKey': package['key'],
        'packageLabel': package['label'],
        'priceCents': package['priceCents'],
        'currency': package['currency'],
        'phrasePool': package['phrasePool'],
        'phone': {
            'number': (phone or {}).get('number'),
            'verified': bool((phone or {}).get('verified')),
        },
        'status': 'draft',
        'paymentMethod': None,
        'paymentMethodLabel': None,
        'transactionLink': None,
        'transactionId': None,
        'createdAt': now,
        'updatedAt': now,
        'messageId': None,
        'approvedAt': None,
        'rejectedAt': None,
        'voucher': None,
    }


def _new_draft(user_id):
    return {
        'userId': str(user_id),
        'step': 'package',
        'package': None,
        'phone': None,
        'method': None,
        'name': None,
        'link': None,
        'txid': None,
        'updatedAt': utcnow(),
    }


def _get_draft(user_id):
    def _read(data):
        drafts = data.setdefault('drafts', {})
        return drafts.get(str(user_id))

    return mutate_json(USERS_FILE, default_users(), _read)


def _save_draft(user_id, draft):
    def _write(data):
        data.setdefault('drafts', {})[str(user_id)] = draft
        return True

    return mutate_json(USERS_FILE, default_users(), _write)


def _submit(data, user_id, draft):
    requests = data.setdefault('requests', {})
    existing = requests.get(str(user_id))
    if existing:
        return existing

    profile = _get_user_profile(data, user_id)
    package = get_package_by_key(draft.get('package')) or DEFAULT_PACKAGES[0]
    phone = draft.get('phone') or {}
    request = _new_request(user_id, profile, package, phone)
    request.update({
        'paymentMethod': draft.get('method'),
        'paymentMethodLabel': (get_payment_method(draft.get('method')) or {}).get('label'),
        'name': draft.get('name'),
        'transactionLink': draft.get('link'),
        'transactionId': draft.get('txid'),
        'status': 'pending',
        'updatedAt': utcnow(),
    })
    requests[str(user_id)] = request
    return request


# ---------------------------------------------------------------------------
# Flow: state machine steps
# ---------------------------------------------------------------------------


def handle_draft_step(user, text):
    """Advance a buyer's draft by one step. Returns (draft, message, reply_markup)."""
    user_id = user.get('id')
    draft = _get_draft(user_id) or _new_draft(user_id)
    step = draft.get('step')
    message = None
    reply_markup = None

    if step == 'package':
        package = get_package_by_key(text)
        if not package:
            return draft, get_package_selection_message(), build_package_keyboard()
        draft.update({'package': package['key'], 'step': 'phone', 'updatedAt': utcnow()})
        _save_draft(user_id, draft)
        message = ('Package: %s\n'
                   'Price: %s\n'
                   'Now share your phone number using the button below.')
        message = message % (package['label'], format_money(package['priceCents'], package['currency']))
        reply_markup = build_phone_share_keyboard()

    elif step == 'phone':
        draft.update({'phone': {'number': text, 'verified': False}, 'step': 'method', 'updatedAt': utcnow()})
        _save_draft(user_id, draft)
        message = 'How do you want to pay?'
        reply_markup = build_payment_method_keyboard()

    elif step == 'method':
        method = get_payment_method(text)
        if not method:
            message = 'Invalid payment method. Use a button below.'
            reply_markup = build_payment_method_keyboard()
            return draft, message, reply_markup
        draft.update({'method': method['key'], 'step': 'name', 'updatedAt': utcnow()})
        _save_draft(user_id, draft)
        message = 'Great. What is your full name?'

    elif step == 'name':
        draft.update({'name': text.strip(), 'step': 'link', 'updatedAt': utcnow()})
        _save_draft(user_id, draft)
        message = ('Send the full transaction link of your payment.\n'
                   'Format: %s') % link_format_hint(draft.get('method'))

    elif step == 'link':
        method = draft.get('method')
        if not is_valid_transaction_link(method, text):
            message = ('That link does not look like a %s receipt.\n'
                       'Expected format: %s\n'
                       'Send the full link again.') % (get_payment_method(method)['label'],
                                                       link_format_hint(method))
            return draft, message, None
        draft.update({'link': text.strip(), 'step': 'txid', 'updatedAt': utcnow()})
        _save_draft(user_id, draft)
        message = ('Now send your transaction ID.\n'
                   'Format: %s') % id_format_hint(method)

    elif step == 'txid':
        method = draft.get('method')
        if not is_valid_transaction_id(method, text):
            message = ('That transaction ID does not look valid for %s.\n'
                       'It should %s.\n'
                       'Try again.') % (get_payment_method(method)['label'], id_format_hint(method))
            return draft, message, None
        draft.update({'txid': text.strip(), 'step': 'done', 'updatedAt': utcnow()})
        _save_draft(user_id, draft)
        return draft, None, None

    return draft, message, reply_markup


def submit_request(user_id, draft):
    def _do(data):
        request = _submit(data, user_id, draft)
        return request

    request = mutate_json(USERS_FILE, default_users(), _do)
    if request.get('status') == 'pending':
        send_message(user_id, build_pending_message(request))
        notify_admin(request)
    return request


def finalize_request(request_id, decision, admin_user):
    """decision: 'approved' or 'rejected'."""
    def _do(data):
        requests = data.get('requests') or {}
        for req in requests.values():
            if req.get('requestId') != request_id:
                continue
            if decision == 'approved':
                req['status'] = 'approved'
                req['approvedAt'] = utcnow()
                req['approvedBy'] = str(admin_user.get('id'))
            else:
                req['status'] = 'rejected'
                req['rejectedAt'] = utcnow()
                req['rejectedBy'] = str(admin_user.get('id'))
            req['updatedAt'] = utcnow()
            return req
        return None

    return mutate_json(USERS_FILE, default_users(), _do)


def assign_voucher(request_id, support_user):
    """Mark a pending_assignment request as delivered and reserve a voucher.
    Returns a status string: 'ok' | 'missing' | 'empty' | 'already'."""
    def _do(data):
        requests = data.get('requests') or {}
        req = None
        for candidate in requests.values():
            if candidate.get('requestId') == request_id:
                req = candidate
                break
        if req is None:
            return 'missing'

        if req.get('voucher'):
            return 'already'

        try:
            entry = reserve_voucher(req.get('phrasePool'))
        except PoolEmptyError:
            req['status'] = 'pending_assignment'
            req['updatedAt'] = utcnow()
            return 'empty'

        owner_profile = _get_user_profile(data, str(req.get('userId')))
        entry.update({
            'ownerId': str(req.get('userId')),
            'ownerName': ('%s %s' % (owner_profile.get('firstName') or '',
                                     owner_profile.get('lastName') or '')).strip(),
            'phone': (req.get('phone') or {}).get('number'),
            'packageKey': req.get('packageKey'),
            'packageLabel': req.get('packageLabel'),
            'assigned': True,
            'status': 'assigned',
            'assignedAt': utcnow(),
            'assignedBy': str(support_user.get('id')),
            'requestId': req.get('requestId'),
        })
        record_voucher_assignment(entry)
        req['voucher'] = {
            'phrase': entry.get('phrase'),
            'extra': entry.get('extra', ''),
            'assignedAt': entry.get('assignedAt'),
            'assignedBy': entry.get('assignedBy'),
            'ownerId': entry.get('ownerId'),
        }
        req['status'] = 'delivered'
        req['updatedAt'] = utcnow()
        return 'ok'

    return mutate_json(USERS_FILE, default_users(), _do)


def send_pending_summary(support_user):
    """Send the list of approved purchases that still need a voucher."""
    def _collect(data):
        pending = []
        for req in (data.get('requests') or {}).values():
            if req.get('status') in ('pending_assignment', 'approved'):
                pending.append(req)
        return pending

    pending = mutate_json(USERS_FILE, default_users(), _collect)
    keyboard = {'inline_keyboard': []}
    for req in pending:
        keyboard['inline_keyboard'].append([{
            'text': 'Assign: %s (%s)' % (req.get('requestId'), req.get('packageLabel')),
            'callback_data': 'assign:%s' % req.get('requestId'),
        }])
    if not keyboard['inline_keyboard']:
        keyboard = None
    send_message(support_user.get('id'), build_pending_assignment_message(pending), reply_markup=keyboard)


# ---------------------------------------------------------------------------
# Update handlers
# ---------------------------------------------------------------------------


def handle_message(update):
    message = update.get('message') or {}
    chat = message.get('chat') or {}
    user = message.get('from') or {}
    chat_id = chat.get('id')
    text = message.get('text') or ''

    if not chat_id:
        return False

    def _record(_data):
        _store_user_profile(_data, user)
        return True

    mutate_json(USERS_FILE, default_users(), _record)

    if text == '/export' and is_admin(chat_id):
        write_exports_and_send(chat_id)
        return True

    if text == '/pending' and can_assign_vouchers(chat_id):
        send_pending_summary(user or {'id': chat_id})
        return True

    if text in ('/start', '/buy'):
        send_message(chat_id, get_start_message(), reply_markup=build_package_keyboard())
        return True

    if text == '/myaccess':
        voucher = get_voucher_for_user(chat_id)
        if voucher:
            status = normalize_voucher_status(voucher)
            package_label = voucher.get('packageLabel') or ''
            if status == 'redeemed':
                send_message(chat_id, 'Your voucher was already redeemed on %s.' % voucher.get('redeemedAt'))
            else:
                send_message(chat_id, 'Your voucher for %s:\nPhrase: %s' % (package_label, voucher.get('phrase')))
        else:
            send_message(chat_id, 'No voucher found. Buy a package first with /buy.')
        return True

    if not text or text.startswith('/'):
        return False

    draft, message, reply_markup = handle_draft_step(user, text)
    if message:
        send_message(chat_id, message, reply_markup=reply_markup)
    elif draft and draft.get('step') == 'done':
        draft['step'] = 'submitted'
        _save_draft(chat_id, draft)
        submit_request(chat_id, draft)
    return True


def handle_contact(update):
    message = update.get('message') or {}
    contact = message.get('contact') or {}
    chat_id = (message.get('chat') or {}).get('id')
    user = message.get('from') or {}

    if not chat_id or not contact:
        return False

    number = contact.get('phone_number') or ''
    if not is_plausible_phone(number):
        send_message(chat_id, 'That phone number looks invalid. Try again.')
        return False

    draft = _get_draft(chat_id) or _new_draft(chat_id)
    if draft.get('step') != 'phone':
        return False

    draft.update({
        'phone': {'number': number, 'verified': True},
        'step': 'method',
        'updatedAt': utcnow(),
    })
    _save_draft(chat_id, draft)
    send_message(chat_id, 'Phone number received. How do you want to pay?',
                 reply_markup=build_payment_method_keyboard())
    return True


def handle_callback(update):
    callback = update.get('callback_query') or {}
    data = callback.get('data') or ''
    user = callback.get('from') or {}
    message = callback.get('message') or {}
    chat_id = (message.get('chat') or {}).get('id')
    callback_id = callback.get('id')

    if not data or not callback_id:
        return False

    def _record(_data):
        _store_user_profile(_data, user)
        return True

    mutate_json(USERS_FILE, default_users(), _record)

    if data.startswith('package:'):
        package_key = data.split(':', 1)[1]
        package = get_package_by_key(package_key)
        if not package:
            answer_callback_query(callback_id, 'Unknown package')
            return True
        draft = _get_draft(user.get('id')) or _new_draft(user.get('id'))
        draft.update({'package': package['key'], 'step': 'phone', 'updatedAt': utcnow()})
        _save_draft(user.get('id'), draft)
        edit_message_reply_markup(chat_id, message.get('message_id'))
        send_message(chat_id,
                     'Package: %s\nPrice: %s\nShare your phone number with the button below.'
                     % (package['label'], format_money(package['priceCents'], package['currency'])),
                     reply_markup=build_phone_share_keyboard())
        return True

    if data.startswith('method:'):
        method_key = data.split(':', 1)[1]
        method = get_payment_method(method_key)
        if not method:
            answer_callback_query(callback_id, 'Unknown method')
            return True
        draft = _get_draft(user.get('id')) or _new_draft(user.get('id'))
        if draft.get('step') != 'method':
            return True
        draft.update({'method': method['key'], 'step': 'name', 'updatedAt': utcnow()})
        _save_draft(user.get('id'), draft)
        edit_message_reply_markup(chat_id, message.get('message_id'))
        send_message(chat_id, 'Great. What is your full name?')
        return True

    if data.startswith('approve:') and is_admin(user.get('id')):
        request_id = data.split(':', 1)[1]
        req = finalize_request(request_id, 'approved', user)
        if not req:
            answer_callback_query(callback_id, 'Request not found')
            return True

        result = assign_voucher(request_id, user)
        if result == 'ok':
            req = get_request_by_id(request_id)
            if req:
                send_message(req.get('userId'), build_approved_message(req.get('packageLabel'),
                                                                     req.get('voucher', {}).get('phrase')))
            answer_callback_query(callback_id, 'Approved and voucher sent')
        elif result == 'already':
            req = get_request_by_id(request_id)
            if req and req.get('voucher'):
                send_message(req.get('userId'), build_approved_message(req.get('packageLabel'),
                                                                     req.get('voucher', {}).get('phrase')))
            answer_callback_query(callback_id, 'Voucher already assigned')
        elif result == 'empty':
            send_message(req.get('userId'), build_approved_pending_voucher_message(req))
            answer_callback_query(callback_id, 'Approved, but voucher pool is empty')
        else:
            answer_callback_query(callback_id, 'Request not found')
        return True

    if data.startswith('reject:') and is_admin(user.get('id')):
        request_id = data.split(':', 1)[1]
        req = finalize_request(request_id, 'rejected', user)
        if req:
            send_message(req.get('userId'), build_rejected_message(req))
        answer_callback_query(callback_id, 'Rejected')
        return True

    if data.startswith('assign:') and can_assign_vouchers(user.get('id')):
        request_id = data.split(':', 1)[1]
        result = assign_voucher(request_id, user)
        if result == 'ok':
            req = get_request_by_id(request_id)
            if req:
                send_message(req.get('userId'), build_approved_message(req.get('packageLabel'),
                                                                     req.get('voucher', {}).get('phrase')))
            answer_callback_query(callback_id, 'Voucher assigned and sent')
            send_pending_summary(user)
        elif result == 'empty':
            answer_callback_query(callback_id, 'Voucher pool is empty')
        elif result == 'already':
            answer_callback_query(callback_id, 'Already assigned')
        else:
            answer_callback_query(callback_id, 'Request not found')
        return True

    return False


def get_request_by_id(request_id):
    data = read_json(USERS_FILE, default_users())
    for req in (data.get('requests') or {}).values():
        if req.get('requestId') == request_id:
            return req
    return None


def write_exports_and_send(chat_id):
    def _collect(data):
        requests = (data.get('requests') or {}).values()
        by_package = {}
        for req in requests:
            by_package.setdefault(req.get('packageKey'), []).append(req)
        return by_package

    by_package = mutate_json(USERS_FILE, default_users(), _collect)
    written = write_exports(by_package)
    for file_path in written:
        send_document(chat_id, file_path)


# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------

app = Flask(__name__)

# PythonAnywhere WSGI entry point imports `application` (see wsgi.py).
application = app


@app.route('/')
def index():
    return jsonify({
        'service': 'tinat-bot',
        'status': 'ok',
        'time': utcnow(),
        'webhook_path': '/telegram-webhook',
        'redeem_path': '/api/v1/vouchers/redeem',
    })


@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'time': utcnow()})


@app.route('/telegram-webhook', methods=['POST'])
def telegram_webhook():
    update = request.get_json(silent=True) or {}
    if not update:
        return '', 200

    update_id = update.get('update_id')
    if record_update(update_id):
        return jsonify({'ok': True, 'duplicate': True})

    try:
        if handle_callback(update) or handle_message(update) or handle_contact(update):
            pass
        return jsonify({'ok': True})
    except Exception:
        logger.exception('Unhandled error while processing update %s', update_id)
        return jsonify({'ok': False, 'error': 'internal error'}), 200


@app.route('/api/v1/vouchers/redeem', methods=['POST'])
def redeem_voucher_api():
    """Redeem a voucher owned by the Telegram user in the Mini App.

    Body: { "initData": "...", "voucherPhrase": "..." }
    The initData hash is verified against the bot token before the voucher is
    atomically marked REDEEMED. A matching entitlement is then recorded.
    """
    if request.content_type not in ('application/json', 'text/json'):
        return jsonify({'error': 'Content-Type must be application/json'}), 415

    body = request.get_json(silent=True) or {}
    init_data = body.get('initData')
    phrase = str(body.get('voucherPhrase') or '').strip()

    if not phrase:
        return jsonify({'error': 'voucherPhrase is required'}), 400

    user = verify_init_data(init_data)
    if user is None:
        return jsonify({'error': 'Invalid or expired initData'}), 401

    owner_id = str(user.get('id'))
    if not rate_limit_redeem(owner_id):
        return jsonify({'error': 'Too many attempts. Try again later.'}), 429

    result = redeem_code(owner_id, phrase)
    if isinstance(result, dict) and 'error' in result:
        code = 404 if result['error'] in ('No matching voucher found', 'Voucher is not assigned to you yet') else 409
        return jsonify(result), code

    return jsonify({'ok': True, 'entitlement': result}), 200


@app.route('/api/v1/vouchers/status', methods=['POST'])
def voucher_status_api():
    """Check whether a voucher is still usable, without consuming it.

    Body: { "initData": "...", "voucherPhrase": "..." }
    """
    if request.content_type not in ('application/json', 'text/json'):
        return jsonify({'error': 'Content-Type must be application/json'}), 415

    body = request.get_json(silent=True) or {}
    init_data = body.get('initData')
    phrase = str(body.get('voucherPhrase') or '').strip()

    if not phrase:
        return jsonify({'error': 'voucherPhrase is required'}), 400

    user = verify_init_data(init_data)
    if user is None:
        return jsonify({'error': 'Invalid or expired initData'}), 401

    voucher = find_voucher(str(user.get('id')), phrase)
    if voucher is None:
        return jsonify({'error': 'No matching voucher found'}), 404

    return jsonify({'ok': True, 'status': normalize_voucher_status(voucher)}), 200


def register_webhook():
    """Register the Telegram webhook for this PythonAnywhere host."""
    if not TELEGRAM_BOT_TOKEN:
        logger.error('Cannot register webhook: missing TELEGRAM_BOT_TOKEN')
        return None

    if 'PA_WEBSITE_URL' in os.environ:
        base_url = os.environ['PA_WEBSITE_URL'].rstrip('/')
    else:
        username = os.environ.get('PA_USERNAME', '').strip()
        base_url = ('https://%s.pythonanywhere.com' % username) if username else 'http://127.0.0.1:5000'

    webhook_url = base_url + '/telegram-webhook'
    allowed = json.dumps(['message', 'callback_query', 'channel_post'])
    payload = {
        'url': webhook_url,
        'allowed_updates': allowed,
        'drop_pending_updates': False,
    }
    result = tg_call('setWebhook', data=payload)
    if result and result.get('ok'):
        logger.info('Webhook registered at %s', webhook_url)
    else:
        logger.error('Failed to register webhook: %s', result)
    return result


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=int(os.environ.get('PORT', '5000')))

